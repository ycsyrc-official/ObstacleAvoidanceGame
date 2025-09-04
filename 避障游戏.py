import pygame
import sys
import random
import math

import os

from openpyxl import Workbook, load_workbook
from datetime import datetime

import tkinter as tk
from tkinter import messagebox



def ui():
    selected_mode = None

    def selected():
        global selected_mode
        selected_mode = var.get()
        root.destroy()


    def information():
        selected_mode = var.get()
        if selected_mode == '普通模式':
            messagebox.showinfo("普通模式", f"{selected_mode}中，你初始会有3条命，每0.2秒会有1%的概率刷新出生命石块（生命+1），撞到石块生命-1，生命为0游戏结束。该模式中，每隔4秒等级提升一次，每级石块速度及生成速度提升5%，最高20级。\n在每轮游戏结束后，游戏记录保存在personal_data_{selected_mode}.xlsx文件中。")
        elif selected_mode == '无限等级模式':
            messagebox.showinfo("无限等级模式", f"{selected_mode}中，你初始会有3条命，每0.2秒会有1%的概率刷新出生命石块（生命+1），撞到石块生命-1，生命为0游戏结束。该模式中，每隔4秒等级提升一次，每级石块速度及生成速度提升5%，没有最高等级限制。\n在每轮游戏结束后，游戏记录保存在personal_data_{selected_mode}.xlsx文件中。")
        elif selected_mode == '生存模式':
            messagebox.showinfo("生存模式", f"{selected_mode}中，你初始会有1条命，且不会刷新出生命石块，撞到石块即游戏结束。该模式中，每隔4秒等级提升一次，每级石块速度及生成速度提升5%，最高20级。\n在每轮游戏结束后，游戏记录保存在personal_data_{selected_mode}.xlsx文件中。")
        elif selected_mode == '特殊模式':
            messagebox.showinfo("特殊模式", f"{selected_mode}中，你初始会有3条命，每0.2秒会有1%的概率刷新出生命石块（生命+1），撞到石块生命-1，生命为0或撞到炸弹（红灰相间的圆）游戏结束。该模式中，每隔4秒等级提升一次，每级石块速度及生成速度提升5%，最高20级。\n该模式中还有两种特殊石块，冰冻石块（蓝色方形）可使所有石块及炸弹速度减半，加速石块（金色方形）可使你控制的小球移速快3倍，两者效果均持续3秒。\n在每轮游戏结束后，游戏记录保存在personal_data_{selected_mode}.xlsx文件中。")


    def on_close():
        # 这里定义点击 × 时的行为，例如弹出确认框
        sys.exit(0)


    # 创建主窗口
    root = tk.Tk()
    root.title("选择游戏模式")
    root.geometry("500x400")

    root.protocol("WM_DELETE_WINDOW", on_close)

    # 模式选项
    modes = ["普通模式", "无限等级模式", "生存模式", "特殊模式"]
    var = tk.StringVar()
    var.set(modes[0])  # 设置默认值

    # 创建选项菜单
    option_menu = tk.OptionMenu(root, var, *modes)
    option_menu.pack(pady=20)
    option_menu.place(x=150, y=50, width=200, height=100)

    # 创建确定按钮
    confirm_button = tk.Button(root, text="确定", command=selected)
    confirm_button.pack()
    confirm_button.place(x=200, y=200, width=100, height=70)

    inform_button = tk.Button(root, text="模式说明", command=information)
    inform_button.pack()
    inform_button.place(x=200, y=300, width=100, height=70)

    # 运行主循环
    root.mainloop()







def game_loop(selected_mode):
    file_path = f"personal_data_{selected_mode}.xlsx"

    # 如果文件不存在就先创建一个
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(row=1, column=1, value='时间')
        ws.cell(row=1, column=2, value='等级')
        ws.cell(row=1, column=3, value='分数')
        wb.save(file_path)

    # 初始化 pygame
    pygame.init()

    # 屏幕尺寸
    SCREEN_WIDTH, SCREEN_HEIGHT = 800, 800
    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))  # 创建屏幕窗口

    pygame.display.set_caption(f"YCSYRC------避障游戏升级版（{selected_mode}）")  # 设置窗口标题

    # 颜色定义
    WHITE = (255, 255, 255)
    BLACK = (0, 0, 0)
    RED = (255, 0, 0)
    DARK_RED = (200, 0, 0)
    GREEN = (0, 255, 0)
    BLUE = (0, 0, 255)
    YELLOW = (255, 255, 0)
    CYAN = (0, 255, 255)
    MAGENTA = (255, 0, 255)
    ORANGE = (255, 165, 0)
    PURPLE = (128, 0, 128)
    PINK = (255, 192, 203)
    GRAY = (128, 128, 128)
    LIGHT_GRAY = (200, 200, 200)
    VERY_LIGHT_GRAY = (240, 240, 240)
    SKY_BLUE = (135, 206, 235)
    LIGHTER_SKY_BLUE = (155, 226, 255)
    GOLD = (255, 215, 0)
    LIGHTER_GOLD = (255, 250, 250)

    stone_color_list = [RED, YELLOW, CYAN, MAGENTA, ORANGE, PURPLE, PINK]

    class Ball:
        def __init__(self, x, y, radius):
            self.x = x
            self.y = y
            self.radius = radius



        def draw(self, surface):
            """绘制物体到指定的屏幕上"""
            pygame.draw.circle(surface, BLACK, (self.x, self.y), self.radius)

        def draw_revive(self, surface, revive_time):
            """绘制物体到指定的屏幕上"""

            if revive_time // 5 % 2 == 0:
                pygame.draw.circle(surface, GRAY, (self.x, self.y), self.radius)
            else:
                pygame.draw.circle(surface, VERY_LIGHT_GRAY, (self.x, self.y), self.radius)




        def move(self, theta, dist, speed_stone):
            if speed_stone == 1:
                if dist < 3:
                    dist = 0
            elif dist < 0.5:
                dist = 0
            """移动板"""
            self.x += math.sqrt(dist) * math.cos(theta) * (speed_stone * 2 + 1)
            self.y += math.sqrt(dist) * math.sin(theta) * (speed_stone * 2 + 1)
            # 确保物体不移出屏幕边界
            self.x = max(0, min(self.x, SCREEN_WIDTH))
            self.y = max(0, min(self.y, SCREEN_HEIGHT))

    class Heart:
        def __init__(self):
            dir = random.randint(0, 3)
            if dir == 0:
                self.x, self.y = 0, random.randint(0, 800)
                self.theta = random.uniform(-math.pi / 2, math.pi / 2)
            if dir == 1:
                self.x, self.y = random.randint(0, 800), 0
                self.theta = random.uniform(0, math.pi)
            if dir == 2:
                self.x, self.y = SCREEN_WIDTH, random.randint(0, 800)
                self.theta = random.uniform(-3 * math.pi / 2, -math.pi / 2)
            if dir == 3:
                self.x, self.y = random.randint(0, 800), SCREEN_HEIGHT
                self.theta = random.uniform(-math.pi, 0)
            self.radius = 20

        def draw(self, surface):

            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), self.radius)
            cx, cy = self.x, self.y
            pts = []
            for i in range(100 + 1):
                t = (i / 100) * 2 * math.pi
                x = 16 * math.sin(t) ** 3
                y = (13 * math.cos(t)
                     - 5 * math.cos(2 * t)
                     - 2 * math.cos(3 * t)
                     - math.cos(4 * t))
                # 翻转 y 轴并缩放、平移
                px = cx + x
                py = cy - y
                pts.append((px, py))
            # 用多边形填充
            pygame.draw.polygon(surface, RED, pts)

        def move(self, ice):
            self.x += 4 * math.cos(self.theta) / (ice + 1)
            self.y += 4 * math.sin(self.theta) / (ice + 1)

    class Stone:
        def __init__(self, radius, color, speed):
            dir = random.randint(0, 3)
            if dir == 0:
                self.x, self.y = 0, random.randint(0, 800)
                self.theta = random.uniform(-math.pi / 2, math.pi / 2)
            if dir == 1:
                self.x, self.y = random.randint(0, 800), 0
                self.theta = random.uniform(0, math.pi)
            if dir == 2:
                self.x, self.y = SCREEN_WIDTH, random.randint(0, 800)
                self.theta = random.uniform(-3 * math.pi / 2, -math.pi / 2)
            if dir == 3:
                self.x, self.y = random.randint(0, 800), SCREEN_HEIGHT
                self.theta = random.uniform(-math.pi, 0)

            self.radius = radius
            self.color = color
            self.speed = speed


        def draw(self, surface):
            """绘制物体到指定的屏幕上"""
            pygame.draw.circle(surface, self.color, (self.x, self.y), self.radius)

        def move(self, ice):

            self.x += self.speed * math.cos(self.theta) / (ice + 1)
            self.y += self.speed * math.sin(self.theta) / (ice + 1)


    class ICE:
        def __init__(self):
            dir = random.randint(0, 3)
            if dir == 0:
                self.x, self.y = 0, random.randint(0, 800)
                self.theta = random.uniform(-math.pi / 2, math.pi / 2)
            if dir == 1:
                self.x, self.y = random.randint(0, 800), 0
                self.theta = random.uniform(0, math.pi)
            if dir == 2:
                self.x, self.y = SCREEN_WIDTH, random.randint(0, 800)
                self.theta = random.uniform(-3 * math.pi / 2, -math.pi / 2)
            if dir == 3:
                self.x, self.y = random.randint(0, 800), SCREEN_HEIGHT
                self.theta = random.uniform(-math.pi, 0)
            self.radius = 20

        def draw(self, surface):

            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), self.radius)
            pygame.draw.rect(surface, SKY_BLUE, (self.x - self.radius / (math.sqrt(2)), self.y - self.radius / (math.sqrt(2)), self.radius * math.sqrt(2), self.radius * math.sqrt(2)))

        def move(self, ice):
            self.x += 4 * math.cos(self.theta) / (ice + 1)
            self.y += 4 * math.sin(self.theta) / (ice + 1)

    class SPEEDSTONE:
        def __init__(self):
            dir = random.randint(0, 3)
            if dir == 0:
                self.x, self.y = 0, random.randint(0, 800)
                self.theta = random.uniform(-math.pi / 2, math.pi / 2)
            if dir == 1:
                self.x, self.y = random.randint(0, 800), 0
                self.theta = random.uniform(0, math.pi)
            if dir == 2:
                self.x, self.y = SCREEN_WIDTH, random.randint(0, 800)
                self.theta = random.uniform(-3 * math.pi / 2, -math.pi / 2)
            if dir == 3:
                self.x, self.y = random.randint(0, 800), SCREEN_HEIGHT
                self.theta = random.uniform(-math.pi, 0)
            self.radius = 20

        def draw(self, surface):

            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), self.radius)
            pygame.draw.rect(surface, GOLD, (self.x - self.radius / (math.sqrt(2)), self.y - self.radius / (math.sqrt(2)), self.radius * math.sqrt(2), self.radius * math.sqrt(2)))

        def move(self, ice):
            self.x += 4 * math.cos(self.theta) / (ice + 1)
            self.y += 4 * math.sin(self.theta) / (ice + 1)

    class BOMBSTONE:
        def __init__(self):
            dir = random.randint(0, 3)
            if dir == 0:
                self.x, self.y = 0, random.randint(0, 800)
                self.theta = random.uniform(-math.pi / 2, math.pi / 2)
            if dir == 1:
                self.x, self.y = random.randint(0, 800), 0
                self.theta = random.uniform(0, math.pi)
            if dir == 2:
                self.x, self.y = SCREEN_WIDTH, random.randint(0, 800)
                self.theta = random.uniform(-3 * math.pi / 2, -math.pi / 2)
            if dir == 3:
                self.x, self.y = random.randint(0, 800), SCREEN_HEIGHT
                self.theta = random.uniform(-math.pi, 0)
            self.radius = 20

        def draw(self, surface):

            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), self.radius)
            pygame.draw.circle(surface, RED, (self.x, self.y), 18)
            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), 16)
            pygame.draw.circle(surface, RED, (self.x, self.y), 14)
            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), 12)
            pygame.draw.circle(surface, RED, (self.x, self.y), 10)
            pygame.draw.circle(surface, LIGHT_GRAY, (self.x, self.y), 8)
            pygame.draw.circle(surface, RED, (self.x, self.y), 6)

        def move(self, ice):
            self.x += 4 * math.cos(self.theta) / (ice + 1)
            self.y += 4 * math.sin(self.theta) / (ice + 1)



    clock = pygame.time.Clock()
    last_time = pygame.time.get_ticks()
    last_time_for_level = pygame.time.get_ticks()
    FPS = 60  # 每秒帧数
    main_ball = Ball(SCREEN_WIDTH / 2, SCREEN_HEIGHT / 2, 20)
    stone_list = []
    heart_list = []
    if selected_mode == '特殊模式':
        ice_list = []
        speed_stone_list = []
        bomb_list = []
    ice_status = 0
    ice_time_cnt = 0
    speed_stone_status = 0
    speed_stone_time_cnt = 0
    button_status = 0
    return_status = 0
    end_status = 0
    score = 0
    level = 1
    if selected_mode == '生存模式':
        life = 1
    else:
        life = 3
    level_update_status = 0
    speed = 4
    speed_mult = 1.05
    revive_time = 100
    image_path = 'YCSYRC.jpg'
    image = pygame.image.load(image_path).convert_alpha()
    image = pygame.transform.smoothscale(image, (100, 100))
    image_rect = pygame.image.load(image_path).convert_alpha().get_rect()
    image_rect.center = (891, 550)

    wb = load_workbook(file_path)
    ws = wb.active

    if len([cell.value for cell in ws['B']]) > 1:
        level_data = [cell.value for cell in ws['B']]
        highest_level = max(level_data[1:])

        score_data = [cell.value for cell in ws['C']]
        highest_score = max(score_data[1:])

    else:
        highest_level = 0
        highest_score = 0


    # 主循环
    running = True
    while running:

        # 事件处理
        for event in pygame.event.get():
            if event.type == pygame.QUIT:  # 关闭窗口事件
                running = False
            elif event.type == pygame.KEYDOWN:  # 按键事件
                if event.key == pygame.K_ESCAPE:  # 按下 ESC 键退出
                    running = False
                if event.key == pygame.K_RETURN:
                    return_status = 1
            elif event.type == pygame.KEYUP:  # 按键事件
                if event.key == pygame.K_RETURN and return_status == 1:
                    last_time = pygame.time.get_ticks()
                    main_ball = Ball(SCREEN_WIDTH / 2, SCREEN_HEIGHT / 2, 20)
                    stone_list = []
                    heart_list = []
                    if selected_mode == '特殊模式':
                        ice_list = []
                        speed_stone_list = []
                        bomb_list = []
                    left_status = 0
                    right_status = 0
                    return_status = 0
                    ice_status = 0
                    ice_time_cnt = 0
                    speed_stone_status = 0
                    speed_stone_time_cnt = 0
                    end_status = 0
                    score = 0
                    level = 1
                    level_update_status = 0
                    speed = 4
                    speed_mult = 1.05
                    if selected_mode == '生存模式':
                        life = 1
                    else:
                        life = 3
                    if len([cell.value for cell in ws['B']]) > 1:
                        level_data = [cell.value for cell in ws['B']]
                        highest_level = max(level_data[1:])

                        score_data = [cell.value for cell in ws['C']]
                        highest_score = max(score_data[1:])

                    else:
                        highest_level = 0
                        highest_score = 0




        if end_status != 1:

            current_time = pygame.time.get_ticks()



            if current_time - last_time >= 200 / (speed / 4):  # 障碍物速度越快，其生成时间间隔越短

                color = random.randint(0, 6)
                new_stone = Stone(20, stone_color_list[color], speed)
                stone_list.append(new_stone)
                last_time = current_time


            if level_update_status == 20:
                speed *= speed_mult
                level += 1
                level_update_status = 0

            if current_time - last_time_for_level >= 200:
                if level < 20 or selected_mode == '无限等级模式':
                    level_update_status += 1
                last_time_for_level = current_time
                if selected_mode == '特殊模式':
                    if random.random() < 0.01:
                        new_ice = ICE()
                        ice_list.append(new_ice)
                    if random.random() < 0.01:
                        speed_stone = SPEEDSTONE()
                        speed_stone_list.append(speed_stone)
                    if random.random() < 0.2:
                        bomb = BOMBSTONE()
                        bomb_list.append(bomb)

                if selected_mode != '生存模式':
                    if random.random() < 0.01:
                        new_heart = Heart()
                        heart_list.append(new_heart)


            mouse_x, mouse_y = pygame.mouse.get_pos()

            dy = mouse_y - main_ball.y
            dx = mouse_x - main_ball.x
            theta = math.atan2(dy, dx)
            dist = math.sqrt(dx ** 2 + dy ** 2)
            main_ball.move(theta, dist, speed_stone_status)

            for stone in stone_list:
                stone.move(ice_status)
            for heart in heart_list:
                heart.move(ice_status)
            if selected_mode == '特殊模式':
                for ice in ice_list:
                    ice.move(ice_status)
                for speed_stone in speed_stone_list:
                    speed_stone.move(ice_status)
                for bomb in bomb_list:
                    bomb.move(ice_status)

            if len(stone_list) != 0:
                for stone_index in range(len(stone_list)):
                    if stone_list[stone_index].x > SCREEN_WIDTH + stone_list[stone_index].radius \
                            or stone_list[stone_index].y > SCREEN_HEIGHT + stone_list[stone_index].radius \
                            or stone_list[stone_index].x < - stone_list[stone_index].radius \
                            or stone_list[stone_index].y < - stone_list[stone_index].radius:
                        stone_list.pop(stone_index)
                        break

            if len(heart_list) != 0:
                for heart_index in range(len(heart_list)):
                    if math.sqrt((heart_list[heart_index].x - main_ball.x) ** 2 + (heart_list[heart_index].y - main_ball.y) ** 2) <= 2 * main_ball.radius:

                        life += 1
                        heart_list.pop(heart_index)
                        break

                for heart_index in range(len(heart_list)):
                    if heart_list[heart_index].x > SCREEN_WIDTH + heart_list[heart_index].radius \
                            or heart_list[heart_index].y > SCREEN_HEIGHT + heart_list[heart_index].radius \
                            or heart_list[heart_index].x < - heart_list[heart_index].radius \
                            or heart_list[heart_index].y < - heart_list[heart_index].radius:
                        heart_list.pop(heart_index)
                        break
            if selected_mode == '特殊模式':
                if len(ice_list) != 0:
                    for ice_index in range(len(ice_list)):
                        if math.sqrt((ice_list[ice_index].x - main_ball.x) ** 2 + (
                                ice_list[ice_index].y - main_ball.y) ** 2) <= 2 * main_ball.radius:
                            ice_status = 1
                            ice_time_cnt = 0
                            ice_list.pop(ice_index)
                            break

                    for ice_index in range(len(ice_list)):
                        if ice_list[ice_index].x > SCREEN_WIDTH + ice_list[ice_index].radius \
                                or ice_list[ice_index].y > SCREEN_HEIGHT + ice_list[ice_index].radius \
                                or ice_list[ice_index].x < - ice_list[ice_index].radius \
                                or ice_list[ice_index].y < - ice_list[ice_index].radius:
                            ice_list.pop(ice_index)
                            break
                if len(speed_stone_list) != 0:
                    for speed_stone_index in range(len(speed_stone_list)):
                        if math.sqrt((speed_stone_list[speed_stone_index].x - main_ball.x) ** 2 + (
                                speed_stone_list[speed_stone_index].y - main_ball.y) ** 2) <= 2 * main_ball.radius:
                            speed_stone_status = 1
                            speed_stone_time_cnt = 0
                            speed_stone_list.pop(speed_stone_index)
                            break

                    for speed_stone_index in range(len(speed_stone_list)):
                        if speed_stone_list[speed_stone_index].x > SCREEN_WIDTH + speed_stone_list[speed_stone_index].radius \
                                or speed_stone_list[speed_stone_index].y > SCREEN_HEIGHT + speed_stone_list[
                            speed_stone_index].radius \
                                or speed_stone_list[speed_stone_index].x < - speed_stone_list[speed_stone_index].radius \
                                or speed_stone_list[speed_stone_index].y < - speed_stone_list[speed_stone_index].radius:
                            speed_stone_list.pop(speed_stone_index)
                            break
                if len(bomb_list) != 0:
                    for bomb_index in range(len(bomb_list)):
                        if bomb_list[bomb_index].x > SCREEN_WIDTH + bomb_list[bomb_index].radius \
                                or bomb_list[bomb_index].y > SCREEN_HEIGHT + bomb_list[
                            bomb_index].radius \
                                or bomb_list[bomb_index].x < - bomb_list[bomb_index].radius \
                                or bomb_list[bomb_index].y < - bomb_list[bomb_index].radius:
                            bomb_list.pop(bomb_index)
                            break



            # 绘制
            if ice_status == 1:
                screen.fill(LIGHTER_SKY_BLUE)  # 用白色填充背景
            elif speed_stone_status == 1:
                screen.fill(LIGHTER_GOLD)
            else:
                screen.fill(WHITE)
            screen.blit(image, image_rect)


            if revive_time >= 100:
                main_ball.draw(screen)
            else:
                main_ball.draw_revive(screen, revive_time)

            for stone in stone_list:
                stone.draw(screen)
            for heart in heart_list:
                heart.draw(screen)
            if selected_mode == '特殊模式':
                for ice in ice_list:
                    ice.draw(screen)
                for speed_stone in speed_stone_list:
                    speed_stone.draw(screen)
                for bomb in bomb_list:
                    bomb.draw(screen)
            pygame.draw.circle(screen, BLACK, (mouse_x, mouse_y), 10, 1)
            font = pygame.font.Font("msyh.ttc", 30)

            text_surface = font.render(f"{selected_mode}", True, BLACK)
            screen.blit(text_surface, (0, 0))

            if score <= highest_score:
                text_surface = font.render(f"分数：{score}({highest_score})", True, BLACK)
            else:
                text_surface = font.render(f"分数：{score}({score})", True, DARK_RED)
            screen.blit(text_surface, (0, 30))
            if level <= highest_level:
                text_surface = font.render(f"等级：{level}({highest_level})", True, BLACK)
            else:
                text_surface = font.render(f"等级：{level}({level})", True, DARK_RED)
            screen.blit(text_surface, (0, 60))
            text_surface = font.render(f"生命：{life}", True, BLACK)
            screen.blit(text_surface, (0, 90))
            if ice_status == 1 and speed_stone_status == 0:
                text_surface = font.render(f"冰冻状态开启", True, BLACK)
                screen.blit(text_surface, (0, 120))
            if speed_stone_status == 1 and ice_status == 0:
                text_surface = font.render(f"小球加速状态开启", True, BLACK)
                screen.blit(text_surface, (0, 120))
            if ice_status == 1 and speed_stone_status == 1:
                text_surface = font.render(f"冰冻状态开启", True, BLACK)
                screen.blit(text_surface, (0, 120))
                text_surface = font.render(f"小球加速状态开启", True, BLACK)
                screen.blit(text_surface, (0, 150))

            text_surface = font.render(f"按回车键重新开始", True, BLACK)
            screen.blit(text_surface, (SCREEN_WIDTH - 240, 0))
            text_surface = font.render(f"按x键返回主页选模式", True, BLACK)
            screen.blit(text_surface, (SCREEN_WIDTH - 285, 30))
            if revive_time <= 100:
                revive_time += 1
            if ice_status == 1:
                if ice_time_cnt <= 180:
                    ice_time_cnt += 1
                else:
                    ice_time_cnt = 0
                    ice_status = 0

            if speed_stone_status == 1:
                if speed_stone_time_cnt <= 180:
                    speed_stone_time_cnt += 1
                else:
                    speed_stone_time_cnt = 0
                    speed_stone_status = 0


            for stone in stone_list:

                if math.sqrt((stone.x - main_ball.x) ** 2 + (stone.y - main_ball.y) ** 2) <= stone.radius + main_ball.radius\
                        and revive_time > 100:

                    if life > 1:
                        life -= 1
                        revive_time = 0
                    else:
                        end_status = 1

                        # 设置字体和大小
                        font = pygame.font.Font("msyh.ttc", 60)

                        text_surface = font.render("游戏结束!", True, BLACK)  # 红色字体
                        screen.blit(text_surface, (SCREEN_WIDTH / 2 - 120, SCREEN_HEIGHT / 2 - 100))  # 将文字绘制到 (50, 100) 坐标
                        text_surface = font.render("按Enter键重新开始", True, BLACK)  # 红色字体
                        screen.blit(text_surface, (SCREEN_WIDTH / 2 - 220, SCREEN_HEIGHT / 2))  # 将文字绘制到 (50, 100) 坐标

                        wb = load_workbook(file_path)
                        ws = wb.active

                        # 找到第一列第一个空单元格的行号
                        row = 1
                        while ws.cell(row=row, column=1).value is not None:
                            row += 1
                        ws.cell(row=row, column=1, value=datetime.now())
                        ws.cell(row=row, column=2, value=level)
                        ws.cell(row=row, column=3, value=score)

                        # 保存为 Excel 文件
                        wb.save(f"personal_data_{selected_mode}.xlsx")
            if selected_mode == '特殊模式':
                for bomb in bomb_list:
                    if math.sqrt((bomb.x - main_ball.x) ** 2 + (bomb.y - main_ball.y) ** 2) <= bomb.radius + main_ball.radius:
                        end_status = 1
                        pygame.draw.circle(screen, RED, (bomb.x, bomb.y), 50)
                        # 设置字体和大小
                        font = pygame.font.Font("msyh.ttc", 60)

                        text_surface = font.render("游戏结束!", True, BLACK)  # 红色字体
                        screen.blit(text_surface, (SCREEN_WIDTH / 2 - 120, SCREEN_HEIGHT / 2 - 100))
                        text_surface = font.render("按Enter键重新开始", True, BLACK)  # 红色字体
                        screen.blit(text_surface, (SCREEN_WIDTH / 2 - 220, SCREEN_HEIGHT / 2))

                        wb = load_workbook(file_path)
                        ws = wb.active

                        # 找到第一列第一个空单元格的行号
                        row = 1
                        while ws.cell(row=row, column=1).value is not None:
                            row += 1
                        ws.cell(row=row, column=1, value=datetime.now())
                        ws.cell(row=row, column=2, value=level)
                        ws.cell(row=row, column=3, value=score)

                        # 保存为 Excel 文件
                        wb.save(f"personal_data_{selected_mode}.xlsx")


            score += 1

        # 刷新屏幕
        pygame.display.flip()

        # 控制帧率
        clock.tick(FPS)


    # 退出 pygame
    pygame.quit()

while True:

    ui()
    game_loop(selected_mode)
